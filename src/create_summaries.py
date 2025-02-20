import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.chart import PieChart, Reference

# Step 1: Load CSV into pandas DataFrame
csv_file = "./output/transactions.csv"  # Replace with your CSV file
excel_file = './output/monthlySummaries.xlsx'
dfRaw = pd.read_csv(csv_file)

#%%#########################
##### FILTER TO MONTHS #####
############################
dfRaw['date'] = dfRaw['createdAt'].apply(lambda x: datetime.strptime(x.split('T')[0],'%Y-%m-%d'))
dfRaw['month'] = dfRaw['date'].apply(lambda x: x.month)
dfRaw['year'] = dfRaw['date'].apply(lambda x: x.year)

startYear = dfRaw['year'].min()
endYear = dfRaw['year'].max()

df = dfRaw
keepCols = ['rawText', 'description', 'message', 'transactionType', 'value', 'category', 'date']
df = df.loc[:,df.columns.isin(keepCols)]
newNames = ['Location', 'Description', 'Message', 'Transaction Type', 'Value', 'Category', 'Date']
df.rename(columns = dict(zip(keepCols, newNames)), inplace = True)
df.index = range(len(df))
# Reorder Columns
df = df[['Date', 'Location', 'Description', 'Message', 'Value', 'Category', 'Transaction Type']]
# Remove transfers to and from savings
df = df.loc[df['Description'] != 'Transfer from Savings',:]
df = df.loc[df['Description'] != 'Transfer to Spending',:]

# Group data by month and sort by most recent dates first
grouped = df.groupby(df['Date'].dt.to_period('M'), group_keys=False)
sorted_groups = sorted(grouped, key=lambda x: x[0], reverse=True)  # Sort by period, descending

# Loop through each month in descending order
for month, group in sorted_groups:
    sheet_name = month.strftime('%Y-%m')  # Format month as "YYYY-MM"

    # Aggregate data for pie chart
    pie_data = group.groupby('Category', as_index=False)['Value'].sum()

    pie_data = pie_data.sort_values(by='Value', ascending=False)
    # Check if Excel file already exists
    try:
        # If file exists, load workbook and add a sheet
        with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            group.to_excel(writer, sheet_name=sheet_name, index=False)
    except FileNotFoundError:
        # If file does not exist, create a new workbook
        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            group.to_excel(writer, sheet_name=sheet_name, index=False)

    # Add the pie chart to the sheet
    workbook = load_workbook(excel_file)
    sheet = workbook[sheet_name]

    # Write pie chart data in the last rows to avoid overwriting
    pie_start_row = len(group) + 5  # Start pie chart data below the main data
    for i, row in pie_data.iterrows():
        sheet.cell(row=pie_start_row + i, column=1, value=row['Category'])
        sheet.cell(row=pie_start_row + i, column=2, value=row['Value'])

    # Create the PieChart
    chart = PieChart()
    chart.title = f"Category Proportion for {sheet_name}"

    # Reference the data and categories
    values = Reference(sheet, min_col=2, min_row=pie_start_row, max_row=pie_start_row + len(pie_data) - 1)
    categories = Reference(sheet, min_col=1, min_row=pie_start_row, max_row=pie_start_row + len(pie_data) - 1)
    chart.add_data(values, titles_from_data=False)
    chart.set_categories(categories)

    # Add the chart to the sheet at K2
    sheet.add_chart(chart, "K2")

    # Save the workbook
    workbook.save(excel_file)

print(f"Data and charts saved to {excel_file}")